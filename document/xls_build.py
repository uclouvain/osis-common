##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2019 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
import datetime
import logging
import re

from django.conf import settings
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook
from openpyxl.styles import Color, Style, PatternFill, Alignment
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.writer.excel import save_virtual_workbook

CONTENT_TYPE_XLS = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; charset=binary'

FIRST_DATA_ROW_NUMBER = 2
FIRST_DATA_COL_NUMBER = 1
MAX_COL_WIDTH = 50
XLS_EXTENSION = 'xlsx'
OPENPYXL_STRING_FORMAT = '@'

LIST_DESCRIPTION_KEY = 'list_description'
FILENAME_KEY = 'filename'
USER_KEY = 'username'

WORKSHEETS_DATA = 'data'
CONTENT_KEY = 'content'
HEADER_TITLES_KEY = 'header_titles'
WORKSHEET_TITLE_KEY = 'worksheet_title'
COLORED_ROWS = 'colored_rows'
COLORED_COLS = 'colored_cols'
STYLED_CELLS = 'styled_cells'
STYLE_NO_GRAY = Style(fill=PatternFill(patternType='solid', fgColor=Color('C1C1C1')))
STYLE_RED = Style(fill=PatternFill(patternType='solid', fgColor=Color(rgb='00FF0000')))
STYLE_BORDER_TOP = Style(
    border=Border(
        top=Side(border_style=BORDER_THIN,
                 color=Color('FF000000')
                 ),
    )
)
STYLE_BORDER_BOTTOM = Style(
    border=Border(
        bottom=Side(border_style=BORDER_THIN,
                    color=Color('FF000000')
                    ),
    )
)

STYLE_MODIFIED = Style(font=Font(color=Color('5CB85C')),)

DESCRIPTION = 'param_description'
FILENAME = 'param_filename'
USER = 'param_user'
HEADER_TITLES = 'param_header_titles'
WS_TITLE = 'param_worksheet_title'

FONT_GREEN = Font(color=Color('5CB85C'))
ROW_HEIGHT = 'row_height'

logger = logging.getLogger(settings.DEFAULT_LOGGER)


def generate_xls(list_parameters, filters=None):
    if _is_valid(list_parameters):
        return _create_xls(list_parameters, filters)
    else:
        logger.warning('Error data invalid to create xls')
        return HttpResponse('')


def _create_xls(parameters_dict, filters=None):
    filename = _build_filename(parameters_dict.get(FILENAME_KEY))

    workbook = Workbook(encoding='utf-8')
    for sheet_number, worksheet_data in enumerate(parameters_dict.get(WORKSHEETS_DATA)):
        _build_worksheet(worksheet_data,  workbook, sheet_number)

    _build_worksheet_parameters(
        workbook,
        parameters_dict.get(USER_KEY),
        parameters_dict.get(LIST_DESCRIPTION_KEY),
        filters
    )
    response = HttpResponse(save_virtual_workbook(workbook), content_type=CONTENT_TYPE_XLS)
    response['Content-Disposition'] = "%s%s" % ("attachment; filename=", filename)

    return response


def _build_worksheet(worksheet_data, workbook, sheet_number):
    content = worksheet_data.get(CONTENT_KEY)

    a_worksheet = _create_worksheet(workbook,
                                    _create_worsheet_title(sheet_number, worksheet_data.get(WORKSHEET_TITLE_KEY)),
                                    sheet_number)
    _add_column_headers(worksheet_data.get(HEADER_TITLES_KEY), a_worksheet)
    _add_content(content, a_worksheet)
    _adjust_column_width(a_worksheet)
    _coloring_rows(a_worksheet, worksheet_data.get(COLORED_ROWS, None))
    _coloring_cols(a_worksheet, worksheet_data.get(COLORED_COLS, None))
    _styling_cells(a_worksheet, worksheet_data.get(STYLED_CELLS, None))
    _format_all_cells_except_header_line(a_worksheet, content)
    if worksheet_data.get(ROW_HEIGHT, None):
        _adjust_row_height(a_worksheet,
                           worksheet_data.get(ROW_HEIGHT).get('height', None),
                           worksheet_data.get(ROW_HEIGHT).get('start', 1),
                           worksheet_data.get(ROW_HEIGHT).get('stop', 1))


def _add_column_headers(headers_title, worksheet1):
    worksheet1.append(_ensure_str_instance(headers_title))


def _add_content(content, a_worksheet_param):
    a_worksheet = a_worksheet_param
    for line in content:
        a_worksheet.append(_validate_fields(line))
    return a_worksheet


def _adjust_column_width(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        for cell in col:
            if cell.coordinate in worksheet.merged_cells:  # not check merge_cells
                continue
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        if adjusted_width > MAX_COL_WIDTH:
            adjusted_width = MAX_COL_WIDTH
            worksheet.column_dimensions[column].width = adjusted_width
        else:
            worksheet.column_dimensions[column].width = adjusted_width


def _build_filename(filename):
    return '%s.%s' % (filename, XLS_EXTENSION)


def _create_worksheet(workbook, title, sheet_num):
    if sheet_num == 0:
        worksheet1 = workbook.active
        worksheet1.title = title
    else:
        worksheet1 = workbook.create_sheet(title=title)

    return worksheet1


def _build_worksheet_parameters(workbook, a_user, list_description=None, filters=None):
    worksheet_parameters = workbook.create_sheet(title=str(_('Parameters')))
    today = datetime.date.today()
    worksheet_parameters.append([str(_('Creation date')), today.strftime('%d-%m-%Y')])
    worksheet_parameters.append([str(_('Created by')), str(a_user)])
    if list_description:
        worksheet_parameters.append([str(_('description')), str(list_description)])
    if filters:
        for key, value in filters.items():
            worksheet_parameters.append([str(key), str(value)])
    _adjust_column_width(worksheet_parameters)
    return worksheet_parameters


def as_text(value):
    if value is None:
        return ""
    return str(value)


def _create_worsheet_title(sheet_number, worksheet_titles):
    current_worksheet_num = sheet_number + 1
    if worksheet_titles:
        return "{} - {}".format(current_worksheet_num, worksheet_titles)

    return "{} - {}{}".format(current_worksheet_num, _('sheet'), current_worksheet_num)


def _is_valid(list_parameters):
    if list_parameters:
        return _is_checked_file_parameters_list(list_parameters) and _is_checked_worsheets_data(list_parameters)
    return False


def _format_all_cells_except_header_line(worksheet1, worksheet_content):
    for row_number, row in enumerate(worksheet_content, FIRST_DATA_ROW_NUMBER):
        for col_number, cell in enumerate(row, FIRST_DATA_COL_NUMBER):
            _adapt_format_for_string_with_numbers(worksheet1, cell, row_number, col_number)
            _align_cells_content(worksheet1, row_number, col_number, horizontal='left', vertical='top')


def _align_cells_content(worksheet1, row_number, col_number, horizontal, vertical):
    worksheet1.cell(column=col_number, row=row_number).alignment = Alignment(horizontal=horizontal,
                                                                             vertical=vertical,
                                                                             wrapText=True)


def _adapt_format_for_string_with_numbers(worksheet1, cell, row_number, col_number):
    """
        Prevent the strings which contains only numbers to be considered as a number and set with a quote while
        looking at the input line
    """
    if type(cell) is str and re.match(r'^[0-9]+$', cell):
        worksheet1.cell(column=col_number, row=row_number).number_format = OPENPYXL_STRING_FORMAT


def _is_checked_file_parameters_list(list_parameters):
    expected_keys = [LIST_DESCRIPTION_KEY,
                     FILENAME_KEY,
                     USER_KEY,
                     WORKSHEETS_DATA]
    return _compare_dicts(expected_keys, list_parameters)


def _compare_dicts(expected_keys, list_parameters):
    keys = list(list_parameters.keys())
    return len((list(set(expected_keys).intersection(set(keys))))) == len(expected_keys)


def _is_checked_worsheets_data(list_parameters):
    expected_keys = [CONTENT_KEY,
                     HEADER_TITLES_KEY,
                     WORKSHEET_TITLE_KEY]
    if list_parameters.get(WORKSHEETS_DATA) is None:
        return False
    else:
        for data in list_parameters.get(WORKSHEETS_DATA):
            if not _compare_dicts(expected_keys, data) or \
                    not _check_correct_number_of_fields(list_parameters.get(HEADER_TITLES_KEY),
                                                        list_parameters.get(CONTENT_KEY)):
                return False
        return True


def _check_correct_number_of_fields(header_titles, content):
    if header_titles and content:
        if len(header_titles) != len(content):
            return False
    return True


def _coloring_rows(ws, data):
    if data:
        for a_style in data.keys():
            row_numbers = data.get(a_style, None)
            if row_numbers:
                _set_row_style(a_style, row_numbers, ws)


def _set_row_style(key, row_numbers, ws):
    for index, row in enumerate(ws.iter_rows()):
        for r in row_numbers:
            if index == r:
                for cell in row:
                    cell.style = key


def _coloring_cols(ws, data):
    if data:
        for a_style in data.keys():
            col_numbers = data.get(a_style)
            if col_numbers:
                _set_col_style(a_style, col_numbers, ws)


def _set_col_style(a_style, col_numbers, ws):
    for index, row in enumerate(ws.iter_rows()):
        for cell in row:
            if cell.col_idx in col_numbers:
                cell.style = a_style


def _validate_fields(line):
    return [as_text(col_content) for col_content in line]


def translate(string_value):
    if type(string_value) == str and string_value:
        return str(_(string_value))
    elif type(string_value) == bool:
        if string_value:
            return _('True')
        return _('False')
    return None


def prepare_xls_parameters_list(working_sheets_data, parameters):
    return {LIST_DESCRIPTION_KEY: _(parameters.get(DESCRIPTION, None)),
            FILENAME_KEY: _(parameters.get(FILENAME, None)),
            USER_KEY: parameters.get(USER, None),
            WORKSHEETS_DATA:
                [{CONTENT_KEY: working_sheets_data,
                  HEADER_TITLES_KEY: parameters.get(HEADER_TITLES, None),
                  WORKSHEET_TITLE_KEY: _(parameters.get(WS_TITLE, None)),
                  STYLED_CELLS: parameters.get(STYLED_CELLS, None),
                  COLORED_ROWS: parameters.get(COLORED_ROWS, None),
                  ROW_HEIGHT: parameters.get(ROW_HEIGHT, None)
                  }
                 ]}


def _styling_cells(ws, data):
    if data:
        for a_style in data.keys():
            cell_reference = data.get(a_style)
            if cell_reference:
                for cell in cell_reference:
                    _set_cell_style(a_style, cell, ws)


def _set_cell_style(a_style, cell_number, ws):
    cell = ws[str(cell_number)]
    ft = a_style
    cell.style = ft


def _adjust_row_height(ws, height, start=1, stop=1):
    if height:
        index = start
        while index <= stop:
            ws.row_dimensions[index].height = height
            index += 1


def _ensure_str_instance(headers_title):
    return [str(title)for title in headers_title]
